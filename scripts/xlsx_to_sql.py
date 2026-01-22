import sys
from openpyxl import load_workbook

# Uso:
# python scripts/xlsx_to_sql.py data/products.xlsx sql/seed_products.sql

inp = sys.argv[1]
outp = sys.argv[2]

wb = load_workbook(inp)
ws = wb.active

headers = [str(c.value).strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
idx = {h: i for i, h in enumerate(headers)}

COL_ID = "ID"
COL_NAME = "TIPO_PRENDA"
COL_DESC = "DESCRIPCIÓN"
COL_STOCK = "CANTIDAD_DISPONIBLE"
COL_AVAILABLE = "DISPONIBLE"
COL_PRICE = "PRECIO_50_U"

required = [COL_ID, COL_NAME, COL_DESC, COL_STOCK, COL_PRICE]
missing = [h for h in required if h not in idx]
if missing:
    raise SystemExit(f"Faltan columnas requeridas: {missing}. Headers: {headers}")

def cell(row, col, default=None):
    i = idx.get(col)
    return row[i].value if i is not None else default

def clean_text(s):
    if s is None:
        return ""
    s = str(s)

    # intenta arreglar mojibake tipo Ã±
    if "Ã" in s or "â" in s:
        try:
            s = s.encode("latin1").decode("utf-8")
        except Exception:
            pass

    fixes = {
        "≤": "ó",
        "±": "ñ",
        "≥": "í",
        "¼": "ü",
        "ß": "á",
        "®": "é",
        "¨": "ú",
        "â€™": "'",
        "â€œ": '"',
        "â€": '"',
        "â€“": "-",
    }
    for k, v in fixes.items():
        s = s.replace(k, v)
    return s

def esc(s):
    return (s or "").replace("'", "''")

def truthy(v):
    if v is None:
        return True
    s = str(v).strip().lower()
    return s in ("1", "true", "si", "sí", "y", "yes", "disponible", "ok")

with open(outp, "w", encoding="utf-8", newline="\n") as f:
    # borrar en orden por FK
    f.write("DELETE FROM cart_items;\n")
    f.write("DELETE FROM carts;\n")
    f.write("DELETE FROM products;\n")

    for r in ws.iter_rows(min_row=2):
        pid = cell(r, COL_ID)
        if pid is None:
            continue
        try:
            pid_int = int(pid)
        except Exception:
            continue

        tipo = clean_text(cell(r, COL_NAME)).strip()
        talla = clean_text(cell(r, "TALLA")).strip()
        color = clean_text(cell(r, "COLOR")).strip()
        categoria = clean_text(cell(r, "CATEGORÍA")).strip()

        name_parts = [tipo]
        if talla: name_parts.append(f"Talla {talla}")
        if color: name_parts.append(color)
        if categoria: name_parts.append(f"({categoria})")
        name = esc(" - ".join([p for p in name_parts if p]))

        desc = esc(clean_text(cell(r, COL_DESC)).strip())

        stock = cell(r, COL_STOCK, 0) or 0
        price = cell(r, COL_PRICE, 0) or 0

        if not truthy(cell(r, COL_AVAILABLE, True)):
            stock = 0

        try:
            stock_int = int(stock)
        except Exception:
            stock_int = 0

        try:
            price_float = float(price)
        except Exception:
            price_float = 0.0

        f.write(
            "INSERT INTO products (id, name, description, price, stock) "
            f"VALUES ({pid_int}, '{name}', '{desc}', {price_float}, {stock_int});\n"
        )

print(f"OK: generado {outp} en UTF-8")
